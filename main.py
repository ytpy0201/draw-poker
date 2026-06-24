import random
import os
from collections import Counter
from datetime import datetime
import openpyxl
from openpyxl import Workbook
import customtkinter as ctk

# 外観モードとテーマの設定
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

# --- ロジック用のクラス定義（poker_v2.0.py から継承） ---
class Card:
    CONVERT = {"J": 11, "Q": 12, "K": 13, "A": 14}
    def __init__(self, suit, rank):
        self.suit = suit
        self.rank = rank
        self.value = self.CONVERT[rank] if rank in self.CONVERT else int(rank)
    def __str__(self):
        return f"{self.suit}{self.rank}"

class Deck:
    def __init__(self):
        suits = ["♠", "♥", "♦", "♣"]
        ranks = ["A", "2", "3", "4", "5", "6", "7", "8", "9", "10", "J", "Q", "K"]
        self.cards = [Card(suit, rank) for suit in suits for rank in ranks]
        self.shuffle()
    def shuffle(self):
        random.shuffle(self.cards)#in-place
    def draw(self):
        if not self.cards:#山札が空なら
            self.__init__()
        return self.cards.pop()

class Player:
    HAND_RANK = {
        "役なし": 0, "ワンペア": 1, "ツーペア": 2, "スリーカード": 3, "ストレート": 4,
        "フラッシュ": 5, "フルハウス": 6, "フォーカード": 7, "ストレートフラッシュ": 8,
        "ロイヤルストレートフラッシュ": 9
    }
    def __init__(self, is_dealer=False, initial_chips=100):
        self.hand = []
        self.is_dealer = is_dealer
        self.chips = initial_chips

    def reset_hand(self):
        self.hand = []

    def draw_initial_hand(self, deck):
        for _ in range(5):
            self.hand.append(deck.draw())

    def judge_hand(self):
        numbers = sorted([card.value for card in self.hand], reverse=True)
        counts = Counter(numbers)
        sorted_counts = sorted(counts.items(), key=lambda x: (x[1], x[0]), reverse=True)#ラムダ式を使った２段階ルール
        compare_list = []
        for num, count in sorted_counts:
            compare_list.extend([num] * count)#リストに別のリストをバラして追加
        count_pattern = sorted(counts.values())

        is_flush = len(set(card.suit for card in self.hand)) == 1
        sorted_nums = sorted(numbers)

        is_straight = (sorted_nums == [2, 3, 4, 5, 14]) or (
            all(sorted_nums[i + 1] - sorted_nums[i] == 1 for i in range(4))#中の条件がすべてTrueのときにTrueを返す
        )
        if is_straight:
            if sorted_nums == [2, 3, 4, 5, 14]: compare_list = [5]
            else: compare_list = [max(numbers)]
        elif is_flush or count_pattern == [1, 1, 1, 1, 1]:
            compare_list = numbers

        is_royal = is_flush and (sorted_nums == [10, 11, 12, 13, 14])

        if is_royal: role = "ロイヤルストレートフラッシュ"
        elif is_flush and is_straight: role = "ストレートフラッシュ"
        elif count_pattern == [1, 4]: role = "フォーカード"
        elif count_pattern == [2, 3]: role = "フルハウス"
        elif is_flush: role = "フラッシュ"
        elif is_straight: role = "ストレート"
        elif count_pattern == [1, 1, 3]: role = "スリーカード"
        elif count_pattern == [1, 2, 2]: role = "ツーペア"
        elif count_pattern == [1, 1, 1, 2]: role = "ワンペア"
        else: role = "役なし"
        return self.HAND_RANK[role], [14] if is_royal else compare_list, role


# --- GUI アプリケーションのメインクラス ---
class PokerApp(ctk.CTk):
    FILENAME = "poker_play_log.xlsx"
    HAND_ODDS = {
        "ロイヤルストレートフラッシュ": 100, "ストレートフラッシュ": 50, "フォーカード": 25,
        "フルハウス": 8, "フラッシュ": 5, "ストレート": 4, "スリーカード": 3,
        "ツーペア": 2, "ワンペア": 1, "役なし": 0
    }

    def __init__(self):
        super().__init__()

        self.title("Python Desktop Poker")
        self.geometry("700x650")
        self.resizable(False, False)

        # ゲーム状態の管理
        self.game_state = "betting" # "betting", "exchanging", "result"
        self.bet = 1
        self.current_winnings = 0  # 今回の勝負での獲得配当を一時保存

        # ゲームデータの初期化
        self.player = Player(is_dealer=False, initial_chips=100)
        self.dealer = Player(is_dealer=True)
        self.deck = Deck()
        self.wb = None
        self.ws = None
        self.setup_excel()

        # --- GUI レイアウトの構築 ---
        self.title_label = ctk.CTkLabel(self, text="♣️ CustomTkinter ポーカー ♣️", font=ctk.CTkFont(size=26, weight="bold"))
        self.title_label.pack(pady=15)

        # ステータス情報エリア
        self.info_frame = ctk.CTkFrame(self)
        self.info_frame.pack(pady=10, fill="x", padx=30)
        self.info_frame.grid_columnconfigure((0, 1), weight=1)

        self.chips_label = ctk.CTkLabel(self.info_frame, text=f"💰 所持チップ: {self.player.chips} 枚", font=ctk.CTkFont(size=18, weight="bold"))
        self.chips_label.grid(row=0, column=0, padx=20, pady=15, sticky="w")

        self.stats_label = ctk.CTkLabel(self.info_frame, text="📊 戦績: 計算中...", font=ctk.CTkFont(size=15))
        self.stats_label.grid(row=0, column=1, padx=20, pady=15, sticky="e")
        self.update_stats_display()

        # ディーラーの手札エリア
        self.dealer_title = ctk.CTkLabel(self, text="🤖 ディーラーの手札", font=ctk.CTkFont(size=14, weight="bold"), text_color="gray")        
        self.dealer_title.pack(pady=(10, 0))
        self.dealer_hand_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.dealer_hand_frame.pack(pady=5)
        self.dealer_card_labels = []
        for i in range(5):
            lbl = ctk.CTkLabel(self.dealer_hand_frame, text="❔", width=80, height=110, fg_color="#2B2B2B", corner_radius=8, font=ctk.CTkFont(size=22))
            lbl.grid(row=0, column=i, padx=8)
            self.dealer_card_labels.append(lbl)

        # プレイヤーの手札エリア
        self.player_title = ctk.CTkLabel(self, text="🧑‍💻 あなたの手札", font=ctk.CTkFont(size=14, weight="bold"), text_color="cyan")        
        self.player_title.pack(pady=(15, 0))
        self.player_hand_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.player_hand_frame.pack(pady=5)
        
        self.player_card_labels = []
        self.checkboxes = []
        
        for i in range(5):
            card_sub_frame = ctk.CTkFrame(self.player_hand_frame, fg_color="transparent")
            card_sub_frame.grid(row=0, column=i, padx=8)

            lbl = ctk.CTkLabel(card_sub_frame, text="❔", width=80, height=110, fg_color="#2B2B2B", corner_radius=8, font=ctk.CTkFont(size=22))
            lbl.pack()
            self.player_card_labels.append(lbl)

            cb = ctk.CTkCheckBox(card_sub_frame, text="交換", font=ctk.CTkFont(size=12))
            cb.pack(pady=8)
            cb.configure(state="disabled")
            self.checkboxes.append(cb)

        # メッセージ・判定結果表示用ラベル
        self.msg_label = ctk.CTkLabel(self, text="ベット枚数を指定して勝負を開始してください。", font=ctk.CTkFont(size=15, weight="bold"))
        self.msg_label.pack(pady=10)

        # 操作コントロールエリア
        self.control_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.control_frame.pack(pady=10)

        self.bet_label = ctk.CTkLabel(self.control_frame, text="ベット数:", font=ctk.CTkFont(size=16))
        self.bet_label.grid(row=0, column=0, padx=10)

        self.bet_entry = ctk.CTkEntry(self.control_frame, width=70, font=ctk.CTkFont(size=16), justify="center")
        self.bet_entry.insert(0, "1")
        self.bet_entry.grid(row=0, column=1, padx=5)

        self.action_button = ctk.CTkButton(self.control_frame, text="勝負を開始する 🃏", command=self.handle_action, font=ctk.CTkFont(size=16, weight="bold"), width=180)
        self.action_button.grid(row=0, column=2, padx=15)

    def setup_excel(self):
        """Excelファイルの読み込み、または新規作成"""
        if os.path.exists(self.FILENAME):
            self.wb = openpyxl.load_workbook(self.FILENAME)
            self.ws = self.wb.active
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "PlayLog"
            self.ws.append(["日時", "ベット数", "プレイヤーの役", "ディーラーの役", "勝敗", "残りチップ"])

    def update_stats_display(self):
        """Excelから勝率を計算して画面に反映"""
        if not os.path.exists(self.FILENAME):
            self.stats_label.configure(text="📊 戦績: データなし")
            return
        try:
            rows = list(self.ws.iter_rows(values_only=True))[1:]
            if not rows:
                self.stats_label.configure(text="📊 戦績: 0回 (0%)")
                return
            total_games = len(rows)
            wins = sum(1 for row in rows if row[4] == "勝ち")
            win_rate = (wins / total_games) * 100 if total_games > 0 else 0
            self.stats_label.configure(text=f"📊 戦績: {total_games}回 (勝率 {win_rate:.1f}%)")
        except:
            self.stats_label.configure(text="📊 戦績: 読み込みエラー")

    def handle_action(self):
        if self.game_state == "betting":
            self.start_round()
        elif self.game_state == "exchanging":
            self.confirm_exchange()
        elif self.game_state == "result":
            self.next_round()

    def start_round(self):
        bet_str = self.bet_entry.get()
        if not bet_str.isdigit():
            self.msg_label.configure(text="❌ 数字を入力してください。", text_color="red")
            return
        
        self.bet = int(bet_str)
        if not 1 <= self.bet <= self.player.chips:
            self.msg_label.configure(text=f"❌ 1 〜 {self.player.chips} の範囲で入力してください。", text_color="red")
            return

        self.player.chips -= self.bet
        self.chips_label.configure(text=f"💰 所持チップ: {self.player.chips} 枚")

        self.deck = Deck()
        self.player.reset_hand()
        self.dealer.reset_hand()
        self.player.draw_initial_hand(self.deck)
        self.dealer.draw_initial_hand(self.deck)

        for i in range(5):
            card = self.player.hand[i]
            color = "#FF4B4B" if card.suit in ["♥", "♦"] else "#FFFFFF"
            self.player_card_labels[i].configure(text=f"{card.suit}\n{card.rank}", text_color=color, fg_color="#3D3D3D")
            self.checkboxes[i].configure(state="normal")
            self.checkboxes[i].deselect()

        for lbl in self.dealer_card_labels:
            lbl.configure(text="❔", text_color="#FFFFFF", fg_color="#2B2B2B")

        self.game_state = "exchanging"
        self.bet_entry.configure(state="disabled")
        self.action_button.configure(text="カードを交換する 🔄")
        self.msg_label.configure(text="交換したいカードにチェックを入れてボタンを押してください。", text_color="white")

    def confirm_exchange(self):
        for i in range(5):
            if self.checkboxes[i].get() == 1:
                self.player.hand[i] = self.deck.draw()
                card = self.player.hand[i]
                color = "#FF4B4B" if card.suit in ["♥", "♦"] else "#FFFFFF"
                self.player_card_labels[i].configure(text=f"{card.suit}\n{card.rank}", text_color=color)
            self.checkboxes[i].configure(state="disabled")

        for i in range(5):
            card = self.dealer.hand[i]
            color = "#FF4B4B" if card.suit in ["♥", "♦"] else "#FFFFFF"
            self.dealer_card_labels[i].configure(text=f"{card.suit}\n{card.rank}", text_color=color, fg_color="#3D3D3D")

        p_score, p_list, p_role = self.player.judge_hand()
        d_score, d_list, d_role = self.dealer.judge_hand()

        player_result_tuple = (p_score, p_list)
        dealer_result_tuple = (d_score, d_list)

        if player_result_tuple > dealer_result_tuple:
            result = "勝ち"
            odds = self.HAND_ODDS[p_role]
            self.current_winnings = self.bet * odds
            
            msg = f"🎉 あなたの【{result}】!! （{p_role} vs {d_role}）\n{self.current_winnings} 枚の配当を獲得！"
            msg_color = "#FFD700"
            
            # 勝利時のみ、ダブルアップ画面を開く誘導メッセージにする
            self.action_button.configure(text="ダブルアップに挑戦 🎰")
        elif player_result_tuple < dealer_result_tuple:
            result = "負け"
            self.current_winnings = 0
            msg = f"💥 あなたの【{result}】... （{p_role} vs {d_role}）\nベットした {self.bet} 枚は没収されました。"
            msg_color = "#FF4B4B"
            self.action_button.configure(text="次のゲームへ 🔁")
        else:
            result = "引き分け"
            self.current_winnings = 0
            self.player.chips += self.bet
            msg = f"🤝【{result}】 （{p_role} vs {d_role}）\nチップが払い戻されます。"
            msg_color = "#FFFFFF"
            self.action_button.configure(text="次のゲームへ 🔁")

        # Excelへログ保存
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.ws.append([current_time, self.bet, p_role, d_role, result, self.player.chips + self.current_winnings])
        self.wb.save(self.FILENAME)

        self.chips_label.configure(text=f"💰 所持チップ: {self.player.chips} 枚")
        self.update_stats_display()
        self.msg_label.configure(text=msg, text_color=msg_color)

        self.game_state = "result"

    def next_round(self):
        """3. 次のゲームへ進む（あるいはダブルアップに入る）判断"""
        if self.current_winnings > 0:
            # 配当がある場合はダブルアップ選択画面を開く
            self.open_double_up_window()
        else:
            self.reset_table_for_next()

    def reset_table_for_next(self):
        """盤面をまっさらにしてベット待ち状態にする"""
        self.game_state = "betting"
        self.bet_entry.configure(state="normal")
        self.action_button.configure(text="勝負を開始する 🃏")
        self.msg_label.configure(text="ベット枚数を指定して勝負を開始してください。", text_color="white")
        
        for i in range(5):
            self.player_card_labels[i].configure(text="❔", fg_color="#2B2B2B", text_color="#FFFFFF")
            self.dealer_card_labels[i].configure(text="❔", fg_color="#2B2B2B", text_color="#FFFFFF")

        if self.player.chips <= 0:
            self.action_button.configure(text="ゲームオーバー ❌")
            self.action_button.configure(state="disabled")
            self.msg_label.configure(text="💥 チップがなくなりました！ゲームオーバーです。", text_color="red")


    # --- 🃏 新機能：ダブルアップ用ポップアップウィンドウ 🃏 ---
    def open_double_up_window(self):
        """別ウィンドウでダブルアップゲームを開く"""
        # 親ウィンドウのボタンを一時的に操作不能にする
        self.action_button.configure(state="disabled")

        # サブウィンドウの作成
        self.du_win = ctk.CTkToplevel(self)
        self.du_win.title("ダブルアップチャンス！")
        self.du_win.geometry("400x450")
        self.du_win.resizable(False, False)
        
        # モーダル（手前のウィンドウを閉じるまで親を触れなくする設定）の再現
        self.du_win.grab_set()

        # サブウィンドウがバツボタンで閉じられたときの処理
        self.du_win.protocol("WM_DELETE_WINDOW", self.close_double_up_quit)

        # 1枚カードを引いて基準にする
        self.parent_card = self.deck.draw()

        # レイアウト
        lbl_title = ctk.CTkLabel(self.du_win, text="🎰 ダブルアップチャンス 🎰", font=ctk.CTkFont(size=20, weight="bold"))
        lbl_title.pack(pady=15)

        self.lbl_pool = ctk.CTkLabel(self.du_win, text=f"現在の権利: {self.current_winnings} 枚", font=ctk.CTkFont(size=16), text_color="yellow")
        self.lbl_pool.pack(pady=5)

        lbl_guide = ctk.CTkLabel(self.du_win, text="次のカードは基準より「大きい」か「小さい」か？", font=ctk.CTkFont(size=13))
        lbl_guide.pack(pady=5)

        # カード表示エリア
        card_frame = ctk.CTkFrame(self.du_win, fg_color="transparent")
        card_frame.pack(pady=15)

        # 基準カード
        p_color = "#FF4B4B" if self.parent_card.suit in ["♥", "♦"] else "#FFFFFF"
        self.lbl_parent = ctk.CTkLabel(card_frame, text=f"{self.parent_card.suit}\n{self.parent_card.rank}", 
                                       width=90, height=130, fg_color="#3D3D3D", corner_radius=8, font=ctk.CTkFont(size=24), text_color=p_color)
        self.lbl_parent.grid(row=0, column=0, padx=20)

        # 次のカード（最初は伏せ）
        self.lbl_child = ctk.CTkLabel(card_frame, text="❔", width=90, height=130, fg_color="#2B2B2B", corner_radius=8, font=ctk.CTkFont(size=24))
        self.lbl_child.grid(row=0, column=1, padx=20)

        # ボタンエリア
        btn_frame = ctk.CTkFrame(self.du_win, fg_color="transparent")
        btn_frame.pack(pady=10)

        self.btn_high = ctk.CTkButton(btn_frame, text="High (大きい)", width=110, command=lambda: self.play_double_up("h"))
        self.btn_high.grid(row=0, column=0, padx=10)

        self.btn_low = ctk.CTkButton(btn_frame, text="Low (小さい)", width=110, command=lambda: self.play_double_up("l"))
        self.btn_low.grid(row=0, column=1, padx=10)

        # 終了してチップ確定ボタン
        self.btn_collect = ctk.CTkButton(self.du_win, text="チップを確定して降りる 💰", fg_color="green", hover_color="darkgreen", command=self.close_double_up_save)
        self.btn_collect.pack(pady=15)

        self.lbl_du_msg = ctk.CTkLabel(self.du_win, text="予想を選んでください。", font=ctk.CTkFont(size=14, weight="bold"))
        self.lbl_du_msg.pack(pady=10)

    def play_double_up(self, choice):
        """HighかLowボタンを押したときのゲーム判定"""
        child_card = self.deck.draw()
        c_color = "#FF4B4B" if child_card.suit in ["♥", "♦"] else "#FFFFFF"
        
        # 開かれたカードを画面に見せる
        self.lbl_child.configure(text=f"{child_card.suit}\n{child_card.rank}", text_color=c_color, fg_color="#3D3D3D")

        if child_card.value == self.parent_card.value:
            self.lbl_du_msg.configure(text="引き分け！カードを引き直します。", text_color="white")
            # 1秒後に自動で次の引き直しに進むなどの処理の代わりに、開いたカードを次の基準にして続行
            self.parent_card = child_card
            self.lbl_parent.configure(text=f"{self.parent_card.suit}\n{self.parent_card.rank}", text_color=c_color)
            self.lbl_child.configure(text="❔", fg_color="#2B2B2B", text_color="#FFFFFF")
            return

        is_high_win = (choice == "h" and child_card.value > self.parent_card.value)
        is_low_win = (choice == "l" and child_card.value < self.parent_card.value)

        if is_high_win or is_low_win:
            self.current_winnings *= 2
            self.lbl_pool.configure(text=f"現在の権利: {self.current_winnings} 枚")
            self.lbl_du_msg.configure(text="🎯 見事的中！配当が2倍になりました！", text_color="gold")
            
            # ボタンを一時無効にして、次のゲームへの「次へ」ボタンに化けさせるか、0.8秒待って次に行く
            # ここではシンプルに、開いたカードを次の基準にして即座に連続挑戦できるようにします
            self.parent_card = child_card
            self.du_win.after(1000, self.prepare_next_double_up)
        else:
            self.current_winnings = 0
            self.lbl_pool.configure(text="現在の権利: 0 枚")
            self.lbl_du_msg.configure(text="❌ 残念！外れました。没収です…", text_color="red")
            
            # 全没収なので操作不能にして終了させる
            self.btn_high.configure(state="disabled")
            self.btn_low.configure(state="disabled")
            self.btn_collect.configure(text="閉じる", fg_color="gray")
            self.current_winnings = 0

    def prepare_next_double_up(self):
        """連勝したときに、開いたカードを左側にずらして次のスタンバイをする"""
        if self.current_winnings == 0:
            return
        p_color = "#FF4B4B" if self.parent_card.suit in ["♥", "♦"] else "#FFFFFF"
        self.lbl_parent.configure(text=f"{self.parent_card.suit}\n{self.parent_card.rank}", text_color=p_color)
        self.lbl_child.configure(text="❔", fg_color="#2B2B2B", text_color="#FFFFFF")
        self.lbl_du_msg.configure(text="次の予想を選んでください。", text_color="white")

    def close_double_up_save(self):
        """チップを確定して終了する"""
        self.player.chips += self.bet + self.current_winnings
        self.current_winnings = 0
        self.chips_label.configure(text=f"💰 所持チップ: {self.player.chips} 枚")
        
        # サブウィンドウを破棄し、メインを復帰
        self.du_win.destroy()
        self.action_button.configure(state="normal")
        self.reset_table_for_next()

    def close_double_up_quit(self):
        """バツボタンで閉じられたり外れて閉じるときの処理"""
        # 途中終了や没収分を反映
        self.player.chips += self.bet + self.current_winnings
        self.current_winnings = 0
        self.chips_label.configure(text=f"💰 所持チップ: {self.player.chips} 枚")
        
        self.du_win.destroy()
        self.action_button.configure(state="normal")
        self.reset_table_for_next()


if __name__ == "__main__":
    app = PokerApp()
    app.mainloop()
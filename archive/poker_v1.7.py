import random
from collections import Counter
from rich.console import Console
from rich.table import Table
from rich.panel import Panel

# --- Excelと日時のためのインポート ---
import os
from datetime import datetime
import openpyxl
from openpyxl import Workbook
# ----------------------------------

console = Console()

# グローバルな定数
HAND_RANK = {
    "役なし": 0,
    "ワンペア": 1,
    "ツーペア": 2,
    "スリーカード": 3,
    "ストレート": 4,
    "フラッシュ": 5,
    "フルハウス": 6,
    "フォーカード": 7,
    "ストレートフラッシュ": 8,
    "ロイヤルストレートフラッシュ": 9
}
HAND_ODDS = {
    "ロイヤルストレートフラッシュ": 100,
    "ストレートフラッシュ": 50,
    "フォーカード": 25,
    "フルハウス": 8,
    "フラッシュ": 5,
    "ストレート": 4,
    "スリーカード": 3,
    "ツーペア": 2,
    "ワンペア": 1,
    "役なし": 0
}


def show_stats(filename="poker_play_log.xlsx"):
    """過去のExcelログを読み込んで、プレイ回数と勝率のみを表示する"""
    if not os.path.exists(filename):
        # まだログファイルがない場合は何も表示せず戻る
        return

    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        
        # 1行目は見出しなので、2行目以降のデータをリスト化
        rows = list(ws.iter_rows(values_only=True))[1:]
        
        if not rows:
            return

        total_games = len(rows)
        wins = 0

        for row in rows:
            # 列の並び： ["日時", "ベット数", "プレイヤーの役", "ディーラーの役", "勝敗", "残りチップ"]
            # 勝敗（インデックス4）のデータをチェック
            result = row[4]
            if result == "勝ち":
                wins += 1

        # 勝率の計算 (小数点第1位まで)
        win_rate = (wins / total_games) * 100 if total_games > 0 else 0

        # richを使って、プレイ回数と勝率をシンプルに表示
        stats_text = (
            f"📊 [bold cyan]通算プレイ回数:[/bold cyan] {total_games} 回\n"
            f"🏆 [bold green]通算勝率      :[/bold green] {win_rate:.1f} % ({wins}勝)"
        )
        
        console.print(Panel(stats_text, title="[bold gold1]🏆 あなたのポーカー戦績 🏆[/bold gold1]", border_style="gold1", expand=False))
        console.print("\n" + "="*40 + "\n")

    except Exception as e:
        # Excelが他のアプリで開かれている場合などのエラー回避
        console.print(f"[dim red]統計データの読み込みに失敗しました: {e}[/dim red]")


def judge_hand(hand):
    numbers = sorted(change_num(hand), reverse=True)
    counts = Counter(numbers)
    
    sorted_counts = sorted(counts.items(), key=lambda x: (x[1], x[0]), reverse=True)
    compare_list = []
    for num, count in sorted_counts:
        compare_list.extend([num] * count)
    
    count_pattern = sorted(counts.values())

    is_royal = royal(hand)
    is_flush = flush(hand)
    is_straight = straight(hand)

    if is_straight:
        if numbers == [14, 5, 4, 3, 2]:
            compare_list = [5]
        else:
            compare_list = [max(numbers)]
    elif is_flush or count_pattern == [1, 1, 1, 1, 1]:
        compare_list = numbers

    if is_royal:
        return HAND_RANK["ロイヤルストレートフラッシュ"], [14]
    elif is_flush and is_straight:
        return HAND_RANK["ストレートフラッシュ"], compare_list
    elif count_pattern == [1, 4]:
        return HAND_RANK["フォーカード"], compare_list
    elif count_pattern == [2, 3]:
        return HAND_RANK["フルハウス"], compare_list
    elif is_flush:
        return HAND_RANK["フラッシュ"], compare_list
    elif is_straight:
        return HAND_RANK["ストレート"], compare_list
    elif count_pattern == [1, 1, 3]:
        return HAND_RANK["スリーカード"], compare_list
    elif count_pattern == [1, 2, 2]:
        return HAND_RANK["ツーペア"], compare_list
    elif count_pattern == [1, 1, 1, 2]:
        return HAND_RANK["ワンペア"], compare_list
    else:
        return HAND_RANK["役なし"], compare_list


def rank_in_hand(hand):
    return [card[1:] for card in hand]


def suit_in_hand(hand):
    return [card[0] for card in hand]


def royal(hand):
    if flush(hand):
        hand = sorted(change_num(hand))
        if hand == [10, 11, 12, 13, 14]:
            return True
    return False


def flush(hand):
    suits = suit_in_hand(hand)
    if len(set(suits)) == 1:
        return True
    else:
        return False


def straight(hand):
    numbers = sorted(change_num(hand))
    if numbers == [2, 3, 4, 5, 14]:
        return True
    for i in range(4):
        if numbers[i + 1] - numbers[i] != 1:
            return False
    return True


def change_num(hand):
    ranks = rank_in_hand(hand)
    convert = {"J": 11, "Q": 12, "K": 13, "A": 14}
    converted = []
    for rank in ranks:
        if rank in convert:
            converted.append(convert[rank])
        else:
            converted.append(int(rank))
    return converted


def exchange_cards(hand, deck):
    console.print(Panel("[yellow]あなたのハンド[/yellow]", expand=False))
    for i, card in enumerate(hand, start=1):
        suit = card[0]
        rank = card[1:]
        color = "red" if suit in ["♥", "♦"] else "white"
        console.print(f"{i}: [{color}]{card}[/{color}]")

    while True:
        choice = input("\n交換したいカードの番号をスペース区切りで入力 \nなければEnter：")
        if choice == "":
            return hand
        
        indexes = choice.split()
        valid = True
        for index in indexes:
            if not index.isdigit() or not 1 <= int(index) <= 5:
                valid = False

        if len(indexes) != len(set(indexes)):
            console.print("[bold red]同じ番号は入力できません[/bold red]")
            continue        

        if not valid:
            console.print("[bold red]１～５の数字を入力してください。[/bold red]")
            continue
        break

    for index in sorted(indexes, reverse=True):
        del hand[int(index) - 1]

    while len(hand) < 5:
        hand.append(deck.pop())

    return hand


def compare_hands(player_result, dealer_result):
    if player_result > dealer_result:
        return "勝ち"
    elif player_result < dealer_result:
        return "負け"
    else:
        return "引き分け"


def make_deck():
    deck = []
    suits = ["♠", "♥", "♦", "♣"]
    ranks = ["A", "2", "3", "4", "5", "6", "7", "8", "9", "10", "J", "Q", "K"]
    for suit in suits:
        for rank in ranks:
            deck.append(f"{suit}{rank}")
    return deck


def deal_cards(deck, hand_list):
    for _ in range(5):
        hand_list.append(deck.pop())


def print_rich_hand(title, hand):
    table = Table(title=title, show_header=False, border_style="bright_blue", box=None)
    for _ in range(len(hand)):
        table.add_column(justify="center", width=6)
        
    row_cells = []
    for card in hand:
        suit = card[0]
        rank = card[1:]
        color = "red" if suit in ["♥", "♦"] else "bright_white"
        card_panel = Panel(f"[{color}]{suit}\n{rank}[/{color}]", border_style="white")
        row_cells.append(card_panel)
        
    table.add_row(*row_cells)
    console.print(table)


def double_up(winnings, deck):
    """ダブルアップゲームを実行する関数"""
    console.print("\n[bold gold1]★★★ ダブルアップチャンス！ ★★★[/bold gold1]")
    while True:
        if len(deck) < 5:
            deck = make_deck()
            random.shuffle(deck)
            
        answer = input(f"現在の獲得チップ {winnings} 枚を賭けてダブルアップに挑戦しますか？ (y/n)：")
        if answer.lower() != "y":
            console.print(f"[green]ダブルアップを終了します。{winnings} 枚のチップを確定しました！[/green]")
            return winnings

        parent_card = deck.pop()
        console.print(Panel(f"基準となるカード: [bold cyan]{parent_card}[/bold cyan]", border_style="cyan", expand=False))
        parent_num = change_num([parent_card])[0]

        while True:
            choice = input("次のカードの数字はこれより大きい？ 小さい？ (h/l)：").lower()
            if choice in ["h", "l"]:
                break
            console.print("[red]h (High) または l (Low) で入力してください。[/red]")

        child_card = deck.pop()
        console.print(Panel(f"開かれたカード: [bold yellow]{child_card}[/bold yellow]", border_style="yellow", expand=False))
        child_num = change_num([child_card])[0]

        if child_num == parent_num:
            console.print("[bold white]数字が同じでした！引き分けです。もう一度同じ枚数で挑戦となります。[/bold white]")
            continue
            
        is_high_win = (choice == "h" and child_num > parent_num)
        is_low_win = (choice == "l" and child_num < parent_num)

        if is_high_win or is_low_win:
            winnings *= 2
            console.print(Panel(f"[bold gold1]見事的中！ 配当が2倍（ {winnings} 枚 ）になりました！[/bold gold1]", border_style="gold1"))
        else:
            console.print(Panel("[bold red]残念！ 予想が外れました。獲得チップは没収です…[/bold red]", border_style="red"))
            return 0


def play_game(player_chips, excel_sheet):
    """引数に excel_sheet を追加し、毎ゲームの記録をできるようにする"""
    console.print(Panel(f"[bold yellow]現在の持ちチップ: {player_chips} 枚[/bold yellow]", border_style="yellow", expand=False))

    while True:
        bet_input = input(f"ベットする枚数を入力してください (1 ～ {player_chips}): ")
        if not bet_input.isdigit():
            console.print("[red]数字を入力してください。[/red]")
            continue
        bet = int(bet_input)
        if not 1 <= bet <= player_chips:
            console.print(f"[red]1 から {player_chips} の範囲で入力してください。[/red]")
            continue
        break

    player_chips -= bet

    deck = make_deck()
    random.shuffle(deck)
    player_hand = []
    dealer_hand = []
    deal_cards(deck, player_hand)
    deal_cards(deck, dealer_hand)

    print_rich_hand("[cyan]◆ 交換前の手札 ◆[/cyan]", player_hand)
    player_hand = exchange_cards(player_hand, deck)

    print_rich_hand("[green]★ あなたの最終手札 ★[/green]", player_hand)
    print_rich_hand("[magenta]♠ ディーラーの手札 ♠[/magenta]", dealer_hand)

    player_result = judge_hand(player_hand)
    dealer_result = judge_hand(dealer_hand)

    player_role_name = [k for k, v in HAND_RANK.items() if v == player_result[0]][0]
    dealer_role_name = [k for k, v in HAND_RANK.items() if v == dealer_result[0]][0]

    console.print(f"\nプレイヤーの役：[bold green]{player_role_name}[/bold green]")
    console.print(f"ディーラーの役：[bold magenta]{dealer_role_name}[/bold magenta]")

    result = compare_hands(player_result, dealer_result)
    
    # 払い戻し用の計算
    if result == "勝ち":
        odds = HAND_ODDS[player_role_name]
        winnings = bet * odds
        console.print(Panel(f"[bold gold1] 判定：{result} !! [/bold gold1]\n役の倍率: {odds}倍 / {winnings} 枚のボーナスを獲得！", border_style="gold1"))
        
        if winnings > 0:
            winnings = double_up(winnings, deck)
            
        player_chips += bet + winnings
    elif result == "引き分け":
        player_chips += bet
        console.print(Panel(f"[bold white] 判定：{result} [/bold white]\nベットしたチップが戻ります。", border_style="white"))
    else:
        console.print(Panel(f"[bold red] 判定：{result} ... [/bold red]\nベットした {bet} 枚のチップが没収されました。", border_style="red"))

    # --- 毎ゲームの結果をExcelシートの行に追記 ---
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    excel_sheet.append([current_time, bet, player_role_name, dealer_role_name, result, player_chips])
    # -------------------------------------------------------------

    return player_chips


if __name__ == "__main__":
    chips = 100
    filename = "poker_play_log.xlsx"

    # --- ゲーム起動時に過去の統計を表示する ---
    show_stats(filename)
    # ----------------------------------------

    # --- メイン開始時にExcelファイルを読み込むか新規作成する ---
    if os.path.exists(filename):
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "PlayLog"
        ws.append(["日時", "ベット数", "プレイヤーの役", "ディーラーの役", "勝敗", "残りチップ"])

    # ゲームのメインループ
    while True:
        # play_gameに関数内部で使うシートオブジェクト(ws)を渡す
        chips = play_game(chips, ws)

        if chips <= 0:
            console.print("\n[bold red]チップがなくなりました！ゲームオーバーです。[/bold red]")
            break

        answer = input("\nもう一回？ (y/n)：")
        if answer.lower() != "y":
            console.print(f"\n最終持ちチップ: [bold yellow]{chips}[/bold yellow] 枚")
            break

    # --- ループを完全に抜けた(ゲーム終了ボタンを押した or 破産した)ときに一括保存 ---
    wb.save(filename)
    console.print(f"[bold green]これまでのプレイログを '{filename}' に保存しました！[/bold green]")
    console.print("Thank you for playing.")
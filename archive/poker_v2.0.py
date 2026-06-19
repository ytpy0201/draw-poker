import random
import os
from collections import Counter

from datetime import datetime
import openpyxl
from openpyxl import Workbook

from rich.console import Console
from rich.table import Table
from rich.panel import Panel

console = Console()


class Card:
    """トランプのカード1枚を表すクラス"""
    CONVERT = {"J": 11, "Q": 12, "K": 13, "A": 14}

    def __init__(self, suit, rank):
        self.suit = suit  # "♠", "♥", "♦", "♣"
        self.rank = rank  # "A", "2", ..., "K"
        # 比較用の数値を自動計算
        self.value = self.CONVERT[rank] if rank in self.CONVERT else int(rank)

    def __str__(self):
        """print(card) としたときに表示される文字列"""
        return f"{self.suit}{self.rank}"


class Deck:
    """52枚の山札を表すクラス"""
    def __init__(self):
        suits = ["♠", "♥", "♦", "♣"]
        ranks = ["A", "2", "3", "4", "5", "6", "7", "8", "9", "10", "J", "Q", "K"]
        self.cards = [Card(suit, rank) for suit in suits for rank in ranks]
        self.shuffle()

    def shuffle(self):
        random.shuffle(self.cards)

    def draw(self):
        """カードを1枚引く"""
        if not self.cards:
            # 万が一山札が切れたら再生成
            self.__init__()
        return self.cards.pop()


class Player:
    """プレイヤー（およびディーラー）を表すクラス"""
    HAND_RANK = {
        "役なし": 0, "ワンペア": 1, "ツーペア": 2, "スリーカード": 3, "ストレート": 4,
        "フラッシュ": 5, "フルハウス": 6, "フォーカード": 7, "ストレートフラッシュ": 8,
        "ロイヤルストレートフラッシュ": 9
    }

    def __init__(self, is_dealer=False, initial_chips=100):
        self.hand = []
        self.is_dealer = is_dealer
        self.chips = initial_chips

    def reset_hand(self):
        self.hand = []

    def draw_initial_hand(self, deck):
        for _ in range(5):
            self.hand.append(deck.draw())

    def judge_hand(self):
        """自身の手札の役とキッカー情報を判定して (役スコア, 比較リスト, 役名) を返す"""
        numbers = sorted([card.value for card in self.hand], reverse=True)
        counts = Counter(numbers)
        
        sorted_counts = sorted(counts.items(), key=lambda x: (x[1], x[0]), reverse=True)
        compare_list = []
        for num, count in sorted_counts:
            compare_list.extend([num] * count)
        
        count_pattern = sorted(counts.values())

        # フラッシュ・ストレートの判定
        is_flush = len(set(card.suit for card in self.hand)) == 1
        
        # ストレート判定（A-5の例外処理含む）
        sorted_nums = sorted(numbers)
        is_straight = (sorted_nums == [2, 3, 4, 5, 14]) or (
            all(sorted_nums[i + 1] - sorted_nums[i] == 1 for i in range(4))
        )

        # キッカーの微調整
        if is_straight:
            if sorted_nums == [2, 3, 4, 5, 14]:
                compare_list = [5]
            else:
                compare_list = [max(numbers)]
        elif is_flush or count_pattern == [1, 1, 1, 1, 1]:
            compare_list = numbers

        # ロイヤル判定
        is_royal = is_flush and (sorted_nums == [10, 11, 12, 13, 14])

        # 役名決定
        if is_royal: role = "ロイヤルストレートフラッシュ"
        elif is_flush and is_straight: role = "ストレートフラッシュ"
        elif count_pattern == [1, 4]: role = "フォーカード"
        elif count_pattern == [2, 3]: role = "フルハウス"
        elif is_flush: role = "フラッシュ"
        elif is_straight: role = "ストレート"
        elif count_pattern == [1, 1, 3]: role = "スリーカード"
        elif count_pattern == [1, 2, 2]: role = "ツーペア"
        elif count_pattern == [1, 1, 1, 2]: role = "ワンペア"
        else: role = "役なし"

        # A-5ストレート、ロイヤルストレートフラッシュの特殊比較スコア
        compare_score = [14] if is_royal else compare_list

        return self.HAND_RANK[role], compare_score, role


class PokerGame:
    """ポーカーゲーム全体の進行・管理を行うクラス"""
    HAND_ODDS = {
        "ロイヤルストレートフラッシュ": 100, "ストレートフラッシュ": 50, "フォーカード": 25,
        "フルハウス": 8, "フラッシュ": 5, "ストレート": 4, "スリーカード": 3,
        "ツーペア": 2, "ワンペア": 1, "役なし": 0
    }
    FILENAME = "poker_play_log.xlsx"

    def __init__(self):
        self.player = Player(is_dealer=False, initial_chips=100)
        self.dealer = Player(is_dealer=True)
        self.deck = None
        self.wb = None
        self.ws = None
        self.setup_excel()

    def setup_excel(self):
        """Excelファイルの読み込み、または新規作成"""
        if os.path.exists(self.FILENAME):
            self.wb = openpyxl.load_workbook(self.FILENAME)
            self.ws = self.wb.active
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "PlayLog"
            self.ws.append(["日時", "ベット数", "プレイヤーの役", "ディーラーの役", "勝敗", "残りチップ"])

    def show_stats(self):
        """過去のExcelログから戦績を表示"""
        if not os.path.exists(self.FILENAME):
            return
        try:
            rows = list(self.ws.iter_rows(values_only=True))[1:]
            if not rows:
                return
            total_games = len(rows)
            wins = sum(1 for row in rows if row[4] == "勝ち")
            win_rate = (wins / total_games) * 100 if total_games > 0 else 0

            stats_text = (
                f"[bold cyan]通算プレイ回数:[/bold cyan] {total_games} 回\n"
                f"[bold green]通算勝率      :[/bold green] {win_rate:.1f} % ({wins}勝)"
            )
            console.print(Panel(stats_text, title="[bold gold1]あなたのポーカー戦績[/bold gold1]", border_style="gold1", expand=False))
            console.print("\n" + "="*40 + "\n")
        except Exception as e:
            console.print(f"[dim red]統計データの読み込みに失敗しました: {e}[/dim red]")

    def print_rich_hand(self, title, player_obj):
        """手札をリッチなカード形式で横並びに表示"""
        table = Table(title=title, show_header=False, border_style="bright_blue", box=None)
        for _ in range(len(player_obj.hand)):
            table.add_column(justify="center", width=6)
            
        row_cells = []
        for card in player_obj.hand:
            color = "red" if card.suit in ["♥", "♦"] else "bright_white"
            card_panel = Panel(f"[{color}]{card.suit}\n{card.rank}[/{color}]", border_style="white")
            row_cells.append(card_panel)
            
        table.add_row(*row_cells)
        console.print(table)

    def exchange_cards(self):
        """プレイヤーのカード交換処理"""
        console.print(Panel("[yellow]あなたのハンド[/yellow]", expand=False))
        for i, card in enumerate(self.player.hand, start=1):
            color = "red" if card.suit in ["♥", "♦"] else "white"
            console.print(f"{i}: [{color}]{card}[/{color}]")

        while True:
            choice = input("\n交換したいカードの番号をスペース区切りで入力 \nなければEnter：")
            if choice == "":
                return
            
            indexes = choice.split()
            valid = True
            for index in indexes:
                if not index.isdigit() or not 1 <= int(index) <= 5:
                    valid = False

            if len(indexes) != len(set(indexes)):
                console.print("[bold red]同じ番号は入力できません[/bold red]")
                continue        
            if not valid:
                console.print("[bold red]１～５の数字を入力してください。[/bold red]")
                continue
            break

        for index in sorted(indexes, reverse=True):
            del self.player.hand[int(index) - 1]

        while len(self.player.hand) < 5:
            self.player.hand.append(self.deck.draw())

    def double_up(self, winnings):
        """ダブルアップゲーム"""
        console.print("\n[bold gold1]ダブルアップチャンス[/bold gold1]")
        while True:
            answer = input(f"現在の獲得チップ {winnings} 枚を賭けてダブルアップに挑戦しますか？ (y/n)：")
            if answer.lower() != "y":
                console.print(f"[green]ダブルアップを終了します。{winnings} 枚のチップを確定しました！[/green]")
                return winnings

            parent_card = self.deck.draw()
            console.print(Panel(f"基準となるカード: [bold cyan]{parent_card}[/bold cyan]", border_style="cyan", expand=False))

            while True:
                choice = input("次のカードの数字はこれより大きい？ 小さい？ (h/l)：").lower()
                if choice in ["h", "l"]:
                    break
                console.print("[red]h (High) または l (Low) で入力してください。[/red]")

            child_card = self.deck.draw()
            console.print(Panel(f"開かれたカード: [bold yellow]{child_card}[/bold yellow]", border_style="yellow", expand=False))

            if child_card.value == parent_card.value:
                console.print("[bold white]数字が同じでした！引き分けです。もう一度同じ枚数で挑戦となります。[/bold white]")
                continue
                
            is_high_win = (choice == "h" and child_card.value > parent_card.value)
            is_low_win = (choice == "l" and child_card.value < parent_card.value)

            if is_high_win or is_low_win:
                winnings *= 2
                console.print(Panel(f"[bold gold1]見事的中！ 配当が2倍（ {winnings} 枚 ）になりました！[/bold gold1]", border_style="gold1"))
            else:
                console.print(Panel("[bold red]残念！ 予想が外れました。獲得チップは没収です…[/bold red]", border_style="red"))
                return 0

    def play_round(self):
        """1ゲーム（ラウンド）の進行"""
        console.print(Panel(f"[bold yellow]現在の持ちチップ: {self.player.chips} 枚[/bold yellow]", border_style="yellow", expand=False))

        while True:
            bet_input = input(f"ベットする枚数を入力してください (1 ～ {self.player.chips}): ")
            if not bet_input.isdigit():
                console.print("[red]数字を入力してください。[/red]")
                continue
            bet = int(bet_input)
            if not 1 <= bet <= self.player.chips:
                console.print(f"[red]1 から {self.player.chips} の範囲で入力してください。[/red]")
                continue
            break

        self.player.chips -= bet

        # デッキと手札の準備
        self.deck = Deck()
        self.player.reset_hand()
        self.dealer.reset_hand()
        self.player.draw_initial_hand(self.deck)
        self.dealer.draw_initial_hand(self.deck)

        self.print_rich_hand("[cyan]◆ 交換前の手札 ◆[/cyan]", self.player)
        self.exchange_cards()

        self.print_rich_hand("[green]★ あなたの最終手札 ★[/green]", self.player)
        self.print_rich_hand("[magenta]♠ ディーラーの手札 ♠[/magenta]", self.dealer)

        # 役判定
        p_score, p_list, p_role = self.player.judge_hand()
        d_score, d_list, d_role = self.dealer.judge_hand()

        console.print(f"\nプレイヤーの役：[bold green]{p_role}[/bold green]")
        console.print(f"ディーラーの役：[bold magenta]{d_role}[/bold magenta]")

        # 勝敗判定 (タプル比較)
        player_result_tuple = (p_score, p_list)
        dealer_result_tuple = (d_score, d_list)

        if player_result_tuple > dealer_result_tuple:
            result = "勝ち"
            odds = self.HAND_ODDS[p_role]
            winnings = bet * odds
            console.print(Panel(f"[bold gold1] 判定：{result} !! [/bold gold1]\n役の倍率: {odds}倍 / {winnings} 枚のボーナスを獲得！", border_style="gold1"))
            if winnings > 0:
                winnings = self.double_up(winnings)
            self.player.chips += bet + winnings
        elif player_result_tuple < dealer_result_tuple:
            result = "負け"
            console.print(Panel(f"[bold red] 判定：{result} ... [/bold red]\nベットした {bet} 枚のチップが没収されました。", border_style="red"))
        else:
            result = "引き分け"
            self.player.chips += bet
            console.print(Panel(f"[bold white] 判定：{result} [/bold white]\nベットしたチップが戻ります。", border_style="white"))

        # Excelログへの追記
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.ws.append([current_time, bet, p_role, d_role, result, self.player.chips])

    def save_and_close(self):
        """Excelファイルを一括保存"""
        self.wb.save(self.FILENAME)
        console.print(f"[bold green]これまでのプレイログを '{self.FILENAME}' に保存しました！[/bold green]")


if __name__ == "__main__":
    # ゲームマネージャーの起動
    game = PokerGame()
    game.show_stats()

    while True:
        game.play_round()

        if game.player.chips <= 0:
            console.print("\n[bold red]チップがなくなりました！ゲームオーバーです。[/bold red]")
            break

        answer = input("\nもう一回？ (y/n)：")
        if answer.lower() != "y":
            console.print(f"\n最終持ちチップ: [bold yellow]{game.player.chips}[/bold yellow] 枚")
            break

    # 終了時に保存
    game.save_and_close()
    console.print("Thank you for playing.")